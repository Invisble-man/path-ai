from __future__ import annotations

import re
from typing import Dict, Any, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


REQ_TRIGGER = re.compile(r"\b(shall|must|will)\b", re.IGNORECASE)
REQ_NUMBERED = re.compile(r"^(\(?[a-z0-9]{1,4}\)?[\.\)]|\d{1,3}\.)\s+", re.IGNORECASE)

ELIGIBILITY_PATTERNS = [
    (re.compile(r"\bservice[-\s]?disabled veteran[-\s]?owned\b|\bsdv?osb\b", re.IGNORECASE), "SDVOSB required"),
    (re.compile(r"\bveteran[-\s]?owned\b|\bvosb\b", re.IGNORECASE), "VOSB required"),
    (re.compile(r"\b8\(a\)\b", re.IGNORECASE), "8(a) required"),
    (re.compile(r"\bhubzone\b", re.IGNORECASE), "HUBZone required"),
    (re.compile(r"\bwosb\b|\bwomen[-\s]?owned\b", re.IGNORECASE), "WOSB required"),
]


def _scan_lines(text: str, max_lines: int = 14000) -> List[str]:
    out = []
    for raw in (text or "").splitlines():
        s = re.sub(r"\s+", " ", raw).strip()
        if s:
            out.append(s)
        if len(out) >= max_lines:
            break
    return out


def extract_requirements_best_effort(rfp_text: str, max_reqs: int = 120) -> List[str]:
    lines = _scan_lines(rfp_text)
    reqs = []
    seen = set()

    for line in lines:
        if len(line) < 25:
            continue

        is_numbered = bool(REQ_NUMBERED.search(line))
        has_trigger = bool(REQ_TRIGGER.search(line))
        if has_trigger and (is_numbered or "offeror" in line.lower() or "proposal" in line.lower() or "submit" in line.lower()):
            key = line.lower().strip()
            if key in seen:
                continue
            seen.add(key)
            reqs.append(line)
            if len(reqs) >= max_reqs:
                break

    return reqs


def classify_requirement(req_text: str) -> Tuple[str, bool, str]:
    t = (req_text or "").lower()

    category = "Technical"
    if "past performance" in t or "cpars" in t:
        category = "Past Performance"
    elif "management" in t or "key personnel" in t or "resume" in t:
        category = "Management"
    elif "price" in t or "pricing" in t or "cost" in t:
        category = "Pricing"
    elif "submit" in t or "deadline" in t or "due" in t or "portal" in t or "email" in t:
        category = "Submission"
    elif "section l" in t or "section m" in t:
        category = "Sections L/M"
    elif "forms" in t or "sf " in t or "dd " in t:
        category = "Forms"

    mandatory = bool(re.search(r"\bshall\b|\bmust\b", t, re.IGNORECASE))

    eligibility_tag = ""
    for pat, tag in ELIGIBILITY_PATTERNS:
        if pat.search(req_text or ""):
            eligibility_tag = tag
            break

    return category, mandatory, eligibility_tag


def _auto_response_section(category: str) -> str:
    mapping = {
        "Submission": "Cover Letter / Admin Volume",
        "Forms": "Admin Volume",
        "Pricing": "Price Volume",
        "Management": "Management Plan",
        "Past Performance": "Past Performance",
        "Sections L/M": "Compliance / Proposal Instructions",
        "Technical": "Technical Approach",
    }
    return mapping.get(category, "Technical Approach")


def build_compatibility_matrix_xlsx(rfp_text: str, company: Dict[str, Any], meta: Dict[str, Any]) -> bytes:
    reqs = extract_requirements_best_effort(rfp_text)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compatibility Matrix"

    # Header block
    ws["A1"] = "Path.ai Compatibility Matrix"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Company: {company.get('legal_name') or '—'}"
    ws["A3"] = f"UEI: {company.get('uei') or '—'}"
    ws["A4"] = f"Solicitation: {meta.get('solicitation') or '—'}"
    ws["A5"] = f"Agency: {meta.get('agency') or '—'}"
    ws["A6"] = f"Due Date: {meta.get('due_date') or '—'}"

    start_row = 8

    headers = [
        "Req ID",
        "Requirement",
        "Category",
        "Mandatory",
        "Eligibility Tag",
        "Response Section",
        "Status",
        "Owner",
        "Notes",
    ]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows
    row = start_row + 1
    for i, req in enumerate(reqs, start=1):
        category, mandatory, eligibility_tag = classify_requirement(req)
        response_section = _auto_response_section(category)

        ws.cell(row=row, column=1, value=f"R{i:03d}")
        ws.cell(row=row, column=2, value=req)
        ws.cell(row=row, column=3, value=category)
        ws.cell(row=row, column=4, value="YES" if mandatory else "NO")
        ws.cell(row=row, column=5, value=eligibility_tag or "")
        ws.cell(row=row, column=6, value=response_section)
        ws.cell(row=row, column=7, value="Unknown")
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value="")

        # Wrap text for requirement + notes
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Column widths
    widths = {
        1: 10,   # Req ID
        2: 85,   # Requirement
        3: 18,   # Category
        4: 10,   # Mandatory
        5: 22,   # Eligibility
        6: 28,   # Response section
        7: 12,   # Status
        8: 14,   # Owner
        9: 34,   # Notes
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze panes at header row
    ws.freeze_panes = ws["A9"]

    # Save
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()