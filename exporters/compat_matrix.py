from __future__ import annotations

import io
from typing import List, Dict, Any


def get_requirements_rows(reqs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for r in (reqs or []):
        rows.append(
            {
                "requirement_id": r.get("requirement_id", ""),
                "requirement": r.get("requirement", ""),
                "status": r.get("status", "Open"),
                "notes": r.get("notes", ""),
            }
        )
    return rows


def build_compatibility_matrix_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compatibility Matrix"

    headers = ["Requirement ID", "Requirement", "Status", "Notes"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in rows or []:
        ws.append(
            [
                r.get("requirement_id", ""),
                r.get("requirement", ""),
                r.get("status", ""),
                r.get("notes", ""),
            ]
        )

    widths = [16, 85, 14, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()